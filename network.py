import collections
import networkconstants as nc
import networkx as nx
import random
from scipy.spatial import distance
import pandas as pd
from openpyxl import load_workbook
import os.path
import Texts_EN as texts


# Generate topology with the specific degrees and weights and the specified number of nodes
def gen_topology(degrees, weights, nodes):
    # Generate k random values with the weight frequencies of degrees
    sequence = random.choices(degrees, weights=weights, k=nodes)
    # The sum of values must be even. If not, increase one of the smallest degree instances
    # print("Sum sequence: ", sum(sequence))
    if (sum(sequence) % 2) != 0:
        # print("Odd sequence")
        min_degree_idx = 0
        min_degree = 1000
        updated = False
        # TODO Keep the smallest degree just in case there is not any with degrees[0]
        for i in range(len(sequence)):
            # Look for the first element with the smallest degree value and increase it, then break
            if sequence[i] == degrees[0]:
                # print("Incremented " + str(i))
                sequence[i] += 1
                updated = True
                break
            elif min_degree > sequence[i]:
                min_degree = sequence[i]
                min_degree_idx = i
        if not updated:
            sequence[min_degree_idx] += 1
            # print("Incremented " + str(min_degree_idx) + " with " + str(min_degree))
    # create the bidirectional network topology using the configuration model
    # it may create parallel edges and self-loops.
    g = nx.configuration_model(sequence)
    # remove parallel edges
    g = nx.Graph(g)
    # Remove self-loops
    g.remove_edges_from(nx.selfloop_edges(g))
    # print(len(g.edges))
    return g


# Calculate the actual frequency of the different degrees in the list of degree_values.
def count_degree_freqs(degree_values):
    counter = collections.Counter(degree_values)
    sorted_values = [counter[x] for x in sorted(counter.keys())]
    frequencies = [100 * val // sum(sorted_values) for val in sorted_values]
    return frequencies


# Calculate distances and scale them to the expected range of values
def calculate_edge_distances(graph, positions, max_limit):
    distances = [distance.euclidean(positions[u], positions[v]) for u, v in graph.edges]
    factor = max_limit / max(distances)
    actual_distances = [int(dist * factor) for dist in distances]
    # print(actual_distances)
    return actual_distances


# Rename nodes combining Type and index
def rename_nodes(assigned_types, types, special_rename_type=None, explicit_values=None, add_prefix=""):
    indexes = dict(zip(types.code, [1] * len(types)))
    # print(indexes)
    # names = [''] * len(assigned_types)
    # for i in range(len(assigned_types)):
    #    names[i] = assigned_types[i] + str(indexes[assigned_types[i]])
    #    indexes[assigned_types[i]] += 1
    # return names
    names, indexes = rename_nodes_idx(assigned_types, indexes, special_rename_type, explicit_values, add_prefix)
    return names


# Rename nodes combining Type and index and receiving initial index per type
def rename_nodes_idx(assigned_types, indexes, special_rename_type=None, explicit_values=None, add_prefix=""):
    names = [''] * len(assigned_types)
    # print(explicit_values)
    for i in range(len(assigned_types)):
        if assigned_types[i] == special_rename_type:
            names[i] = explicit_values[indexes[assigned_types[i]] - 1]
        else:
            names[i] = assigned_types[i] + add_prefix + str(indexes[assigned_types[i]])
        indexes[assigned_types[i]] += 1
    return names, indexes


# Write Excel file with two sheets
def write_network_xls(file, graph, distances, types, node_sheet, link_sheet, macro_region=None, x_coord=None,
                      y_coord=None, reference_nodes=None, coord_type=nc.BACKBONE, households=None, macro_cells=None,
                      small_cells=None, regional_twins=None, national_twins=None, link_capacities=None):
    x_coord_back = [0.00] * len(graph.nodes)
    y_coord_back = [0.00] * len(graph.nodes)
    x_coord_m_core = [0.00] * len(graph.nodes)
    y_coord_m_core = [0.00] * len(graph.nodes)

    match coord_type:
        case nc.METRO_CORE:
            x_coord_m_core = x_coord
            y_coord_m_core = y_coord
        case _:
            x_coord_back = x_coord
            y_coord_back = y_coord
            if reference_nodes is None:
                reference_nodes = [val for val in graph.nodes]

    if macro_region is None:
        macro_region = [0] * len(graph.nodes)

    if reference_nodes is None:
        reference_nodes = [''] * len(graph.nodes)

    if national_twins is None:
        national_twins = [''] * len(graph.nodes)

    if regional_twins is None:
        regional_twins = [''] * len(graph.nodes)

    if households is None:
        households = [12000] * len(graph.nodes)

    if macro_cells is None:
        macro_cells = [8] * len(graph.nodes)

    if small_cells is None:
        small_cells = [24] * len(graph.nodes)

    if link_capacities is None:
        link_capacities = [100] * len(graph.edges)

    nodes_df = pd.DataFrame({nc.XLS_NODE_NAME: [val for val in graph.nodes],
                             nc.XLS_CO_TYPE: types,
                             nc.XLS_REF_RCO: [val for val in graph.nodes],
                             nc.XLS_REF_NCO: reference_nodes,
                             nc.XLS_HOUSE_H: households,
                             nc.XLS_MACRO_C: macro_cells,
                             nc.XLS_SMALL_C: small_cells,
                             nc.XLS_TWIN_RCO: regional_twins,
                             nc.XLS_TWIN_NCO: national_twins,
                             nc.XLS_CLUSTER: macro_region,
                             nc.XLS_X_BACK: x_coord_back,
                             nc.XLS_Y_BACK: y_coord_back,
                             nc.XLS_X_MCORE: x_coord_m_core,
                             nc.XLS_Y_MCORE: y_coord_m_core,
                             })
    links_df = pd.DataFrame({nc.XLS_SOURCE_ID: [u for u, v in graph.edges],
                             nc.XLS_DEST_ID: [v for u, v in graph.edges],
                             nc.XLS_DISTANCE: distances,
                             nc.XLS_CAPACITY_GBPS: link_capacities}
                            )
    if os.path.isfile(file):
        res, message, nodes_df, links_df = merge_data_frames_to_xls(file, node_sheet, link_sheet, nodes_df, links_df, coord_type)
        if not res:
            return res, message
        with pd.ExcelWriter(file, engine='openpyxl') as writer:
            nodes_df.to_excel(writer, sheet_name=node_sheet, index=False)
            links_df.to_excel(writer, sheet_name=link_sheet, index=False)
            # with pd.ExcelWriter(file, engine='openpyxl', mode="a", if_sheet_exists='overlay') as writer:
            '''book = load_workbook(file)
            # writer.book=book
            writer.sheets.update(dict((ws.title, ws) for ws in book.worksheets))'''

            '''nodes_df.to_excel(writer, sheet_name=node_sheet, index=False, header=False,
                              startrow=writer.sheets[node_sheet].max_row)'''
            '''links_df.to_excel(writer, sheet_name=link_sheet, index=False, header=False,
                              startrow=writer.sheets[link_sheet].max_row)'''
    else:
        with pd.ExcelWriter(file, engine='openpyxl') as writer:
            nodes_df.to_excel(writer, sheet_name=node_sheet, index=False)
            links_df.to_excel(writer, sheet_name=link_sheet, index=False)

    return True, ""


def merge_data_frames_to_xls(file, node_sheet, link_sheet, nodes_df, links_df, coord_type):
    ex_nodes = pd.read_excel(file, node_sheet, engine="openpyxl")
    existing_links = pd.read_excel(file, link_sheet, engine="openpyxl")
    node_names = nodes_df[nc.XLS_NODE_NAME]
    same = pd.merge(node_names, ex_nodes, on=[nc.XLS_NODE_NAME])
    link_ids = existing_links[[nc.XLS_SOURCE_ID, nc.XLS_DEST_ID]]
    same_links = pd.merge(link_ids, links_df, on=[nc.XLS_SOURCE_ID, nc.XLS_DEST_ID])

    if len(same_links) > 0:
        return False, texts.LINK_EXISTS, nodes_df

    # We are generating a backbone network, but there is already one defined
    if coord_type == nc.BACKBONE and max(ex_nodes[nc.XLS_X_BACK]) > 0:
        return False, texts.BACKBONE_EXISTS, nodes_df

    # We are generating a metro core network but the node was already used for another macro region
    if coord_type == nc.METRO_CORE and max(same[nc.XLS_X_MCORE]) > 0:
        return False, texts.METRO_CORE_CONFLICT, nodes_df, links_df

    # We are creating a metro core network and one of the nodes already exists (probably mapped from
    # the backbone network. Let's add information related to the coordinates and update other fields
    if coord_type == nc.METRO_CORE and len(same) > 0:
       #nc.XLS_X_MCORE, nc.XLS_Y_MCORE]
        for name in same[nc.XLS_NODE_NAME]:
            # TODO read all at once and write them together
            hh = nodes_df.loc[nodes_df[nc.XLS_NODE_NAME] == name, nc.XLS_HOUSE_H].iloc[0]
            mc = nodes_df.loc[nodes_df[nc.XLS_NODE_NAME] == name, nc.XLS_MACRO_C].iloc[0]
            sc = nodes_df.loc[nodes_df[nc.XLS_NODE_NAME] == name, nc.XLS_SMALL_C].iloc[0]
            x_coord = nodes_df.loc[nodes_df[nc.XLS_NODE_NAME] == name, nc.XLS_X_MCORE].iloc[0]
            y_coord = nodes_df.loc[nodes_df[nc.XLS_NODE_NAME] == name, nc.XLS_Y_MCORE].iloc[0]

            ex_nodes.loc[ex_nodes[nc.XLS_NODE_NAME] == name, nc.XLS_HOUSE_H] = hh
            ex_nodes.loc[ex_nodes[nc.XLS_NODE_NAME] == name, nc.XLS_MACRO_C] = mc
            ex_nodes.loc[ex_nodes[nc.XLS_NODE_NAME] == name, nc.XLS_SMALL_C] = sc
            ex_nodes.loc[ex_nodes[nc.XLS_NODE_NAME] == name, nc.XLS_X_MCORE] = x_coord
            ex_nodes.loc[ex_nodes[nc.XLS_NODE_NAME] == name, nc.XLS_Y_MCORE] = y_coord

            nodes_df = nodes_df[nodes_df[nc.XLS_NODE_NAME] != name]

        ex_nodes.update(ex_nodes, overwrite=True)

    nodes_df = pd.concat([ex_nodes, nodes_df])
    links_df = pd.concat([existing_links, links_df])
    return True, texts.NO_ERROR, nodes_df, links_df


# Write Excel file with two sheets
def write_network_xls_plus(file, graph, distances, types, node_sheet, link_sheet, reference_nodes):
    nodes_df = pd.DataFrame({'Nodes': [val for val in graph.nodes],
                             'Degrees': [val for u, val in graph.degree],
                             'Type': types,
                             'Reference': reference_nodes})
    links_df = pd.DataFrame({'Node1': [u for u, v in graph.edges],
                             'Node2': [v for u, v in graph.edges],
                             'Distance': distances}
                            )
    with pd.ExcelWriter(file, engine='openpyxl') as writer:
        nodes_df.to_excel(writer, sheet_name=node_sheet, index=False)
        links_df.to_excel(writer, sheet_name=link_sheet, index=False)
    return True


# Extract proportion of distances in each defined range
def count_distance_ranges(distances, upper_limits):
    occurrences = [0] * len(upper_limits)
    for dist in distances:
        i = 0
        for limit in upper_limits:
            if dist < limit:
                occurrences[i] += 1
                break
            i += 1
    percentages = [round(100 * occurrence / sum(occurrences), 2) for occurrence in occurrences]
    return percentages


# Assign types to graph nodes based on type proportions
def backbone_assign_types(graph, types):
    # Generate k random values with the weight frequencies of degrees
    sequence = random.choices(types.code, weights=types.proportion, k=len(graph.nodes))
    return sequence


# Assign types to graph nodes based on type proportions
def metro_assign_types(types):
    # print("metro_assign_types: ", types)
    sequence = [code for code, val in types[['code', 'number']].values for _ in range(val)]
    # Generate k random values with the weight frequencies of degrees
    random.shuffle(sequence)
    return sequence


def set_missing_types_colors(assigned_types, dict_colors: dict):
    unique_types = set(assigned_types)
    for a_type in unique_types:
        color = dict_colors.get(a_type, None)
        if color is None:
            dict_colors[a_type] = 'b'
    return dict_colors


def color_nodes(assigned_types, dict_colors):
    dict_colors = set_missing_types_colors(assigned_types, dict_colors)
    return [dict_colors[assigned_type] for assigned_type in assigned_types]


# TODO: Matching types, but not checking degrees.
# Strategy 1: Relabel 2 nodes to match other existing ones with the same type
# See also connect_sub_metro_regions
def match_sub_metro_region(topo_nodes, assigned_types, sub_topo_nodes, sub_topo_types):
    # relabelled = [val + len(topo_nodes)-2 for val in sub_topo_nodes]
    relabelled = [val + len(topo_nodes) for val in sub_topo_nodes]
    type_1 = sub_topo_types[0]
    idx_1 = assigned_types.index(type_1)
    type_2 = sub_topo_types[1]
    idx_2 = assigned_types.index(type_2, idx_1 + 1)
    result = [idx_1, idx_2]
    result.extend(relabelled[2:(len(relabelled))])
    # print(result)
    return result, sub_topo_types[2:]


# Strategy 2: Connect nodes to other existing ones. Increase degrees.
# See also match_sub_metro_regions
def connect_sub_metro_regions(topology, sub_topology, min_degree):
    indexes = []
    low_degree = min_degree
    while True:
        temp = [index for index, element in topology.degree if element <= low_degree]
        indexes.extend(temp)
        if len(indexes) > 1:
            break
        low_degree += 1
    topo_cons = random.sample(indexes, k=2)
    relabelled = [val + len(topology.nodes) for val in sub_topology.nodes]

    indexes = []
    while True:
        temp = [index + len(topology.nodes) for index, element in sub_topology.degree if element == low_degree]
        indexes.extend(temp)
        if len(indexes) > 1:
            break
        low_degree += 1
    subtopo_cons = random.sample(indexes, k=2)

    sub_topology = nx.relabel_nodes(sub_topology, dict(zip(sub_topology.nodes, relabelled)))
    topology = nx.union(topology, sub_topology)
    topology.add_edge(topo_cons[0], subtopo_cons[0])
    topology.add_edge(topo_cons[1], subtopo_cons[1])
    return topology


def format_distance_limits(distances, upper_limits):
    init_dist = 0
    text = "% of distances per range (kms):\n"
    for upper, perc in zip(upper_limits, (count_distance_ranges(distances, upper_limits))):
        text += (str(init_dist) + "-" + str(upper) + ":\t" + str(perc) + "%\n")
        init_dist = upper
    return text
